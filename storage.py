from subprocess import Popen, PIPE

import pickle
import datetime
import os


# http://macappstore.org/antiword/
DATE_FORMAT = '%d%m%y-%H%M'


def convert_pdf_to_txt(path):
    from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
    from pdfminer.converter import TextConverter
    from pdfminer.layout import LAParams
    from pdfminer.pdfpage import PDFPage
    from cStringIO import StringIO
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = file(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos = set()
    for page in PDFPage.get_pages(fp, pagenos,
                                  maxpages=maxpages, password=password,
                                  caching=caching, check_extractable=True):
        interpreter.process_page(page)
    fp.close()
    device.close()
    str = retstr.getvalue()
    retstr.close()
    return str


def convert_xls_to_txt(file_path, ext):
    import xlrd
    from openpyxl import load_workbook
    if ext == 'xls':
        wb = xlrd.open_workbook(file_path)
        sheets = wb.sheets()
        title = 'name'
        rows = 'get_rows'
    else:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.worksheets
        title = 'title'
        rows = 'iter_rows'
    data = []
    for sh in sheets:
        sh_title = getattr(sh, title)
        sh_data = ['title:{}\n'.format(sh_title)]
        for r in getattr(sh, rows)():
            # joiner = '\t;'
            joiner = '\n'
            r_data = []
            for c in r:
                val = c.value or ''
                if isinstance(val, datetime.datetime):
                    val = val.strftime(DATE_FORMAT)
                if not isinstance(val, basestring):
                    val = str(val)
                r_data.append(val.encode('ascii', errors='ignore'))
            sh_data.append(joiner.join(r_data))
        sh_data = '\n'.join(sh_data)
        data.append(sh_data)
    return '--sheet--\n'.join(data)


def convert_doc_to_txt(file_path):
    from docx import opendocx, getdocumenttext
    document = opendocx(file_path)
    paratextlist = getdocumenttext(document)
    newparatextlist = []
    for paratext in paratextlist:
        newparatextlist.append(paratext.encode("utf-8"))
    return '\n\n'.join(newparatextlist)


def read_file(filename, file_path):
    # http://davidmburke.com/2014/02/04/python-convert-documents-doc-docx-odt-pdf-to-plain-text-without-libreoffice/
    ext = filename.split('.')[-1]
    if ext in ['xlsx', 'xls']:
        return convert_xls_to_txt(file_path, ext)
    elif ext == "doc":
        cmd = ['antiword', file_path]
        p = Popen(cmd, stdout=PIPE)
        stdout, stderr = p.communicate()
        return stdout.decode('ascii', 'ignore')
    elif ext == "docx":
        return convert_doc_to_txt(file_path)
    elif ext == "odt":
        cmd = ['odt2txt', file_path]
        p = Popen(cmd, stdout=PIPE)
        stdout, stderr = p.communicate()
        return stdout.decode('ascii', 'ignore')
    elif ext == "pdf":
        return convert_pdf_to_txt(file_path)
    else:
        f = open(file_path, 'r+')
        text = f.read()
        f.close()
        return text


class Storage():

    def __str__(self):
        return self.STORAGE

    def __init__(self, dir_name, parent_folder=None):
        self.EXT = dir_name
        self.STORAGE, exists = self.join(dir_name, parent_folder, create=True)

    def create_sub_storage(self, subfolder, date=True):
        parent = self.STORAGE
        # if date:
        #     parent, exists = self.join(subfolder, create=True)
        #     subfolder = datetime.datetime.now().strftime('%d%m%y-%H%M')
        return Storage(subfolder, parent)

    def join(self, subfolder, folder=None, create=False):
        if folder is None:
            folder = self.STORAGE
        path = os.path.join(folder, subfolder)
        exists = False
        if not os.path.exists(path):
            if create:
                os.makedirs(path)
        else:
            exists = True
        return path, exists

    def generate_name(self, name='all', ftype=''):
        ext = self.EXT
        date = datetime.datetime.now().strftime(DATE_FORMAT)
        return '{ftype}_{name}_{date}.c{ext}'.format(
            ftype=ftype, name=name, date=date, ext=ext)

    def write(self, text, name='all', ftype=None, rewrite=True, concat=False):
        def write(file_path, text, line_mode='ml'):
            if line_mode == 'sl':
                text = text.replace('\n', ' ')
            if ftype:
                file_path = file_path + line_mode
            f = open(file_path, 'w+')
            f.write(text)
            f.close()
        if name.startswith(ftype):
            name = name.split('_')[1]
        name = name[:10]
        if ftype:
            name = self.generate_name(name, ftype)
        file_path, exists = self.join(name)
        if not rewrite and exists:
            raise ValueError('File {} already exists'.format(file_path))
        write(file_path, text)
        if concat:
            write(file_path, text, 'sl')
        # print('Write file {}'.format(file_path))
        return file_path

    def write_pickle(self, data, name='all', ftype=None, ext=''):
        name = self.generate_name(name, ext + 'meta', ftype)
        file_path, exists = self.join(name)
        f = open(file_path, 'wb')
        pickle.dump(data, f)
        f.close()
        return file_path

    def read_pickle(self, name='all', ftype=None, ext=None):
        name = self.generate_name(name, ext, ftype)
        file_path, exists = self.join(name)
        f = open(file_path, 'rb')
        data = pickle.load(f)
        f.close()
        return data

    def read(self, name):
        file_path, exists = self.join(name, create=False)
        if not exists:
            raise ValueError('File {} is not exist'.format(file_path))
        return read_file(name, file_path)

    def files(self, filtype=None):
        def get_date(filename):
            filename = filename.split('.')[0]
            date = filename.split('_')[-1]
            return datetime.datetime.strptime(date, DATE_FORMAT)
        files = []
        for f in os.listdir(self.STORAGE):
            if os.path.isfile(os.path.join(self.STORAGE, f)):
                if filtype and not f.startswith(filtype):
                    continue
                if f.startswith('.~lock'):
                    continue
                files.append(f)
        if filtype:
            return sorted(files, key=get_date, reverse=True)
        return files
